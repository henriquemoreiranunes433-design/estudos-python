from pathlib import Path

import openpyxl

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

estudantes = [
    
    ['João', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10],
]

# for i, linha_estudantes in enumerate(estudantes, start=2):
#     for j, cooluna_estudante in enumerate(linha_estudantes, start=1):
#         print(i, j, cooluna_estudante)
#         worksheet.cell(i, j, cooluna_estudante)
for estudante in estudantes:
    worksheet.append(estudante)


workbook.save(WORKBOOK_PATH)